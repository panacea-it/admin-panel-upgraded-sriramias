"""Generate Finance Module Frontend Audit Report (Excel)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

OUTPUT = "Finance-Module-Frontend-Audit-Report.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1A3A5C")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor="246392")
SUBHEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SECTION_FILL = PatternFill("solid", fgColor="EEF6FC")
SECTION_FONT = Font(bold=True, size=11, color="1A3A5C")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
        cell.border = BORDER


def style_data_rows(ws, start_row, end_row, cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.border = BORDER


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_table(ws, start_row, headers, rows):
    cols = len(headers)
    for c, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=c, value=h)
    style_header_row(ws, start_row, cols)
    r = start_row + 1
    for row in rows:
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
        r += 1
    style_data_rows(ws, start_row + 1, r - 1, cols)
    return r


def add_module_sheet(wb, title, module_data):
    ws = wb.create_sheet(title[:31])
    ws["A1"] = module_data["name"]
    ws["A1"].font = Font(bold=True, size=14, color="1A3A5C")
    ws.merge_cells("A1:F1")

    r = 3
    sections = [
        ("Route / Component", module_data.get("route", "")),
        ("Page Component", module_data.get("component", "")),
        ("Completion %", module_data.get("completion", "")),
        ("Implementation Status", module_data.get("status", "")),
        ("Data Source", module_data.get("data_source", "")),
    ]
    for label, val in sections:
        ws.cell(row=r, column=1, value=label).font = SECTION_FONT
        ws.cell(row=r, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        ws.cell(row=r, column=2, value=val)
        r += 1

    r += 1
    for section_title, items in [
        ("Existing Features (Visible UI)", module_data.get("existing", [])),
        ("Implemented Features", module_data.get("implemented", [])),
        ("Partial Implementations", module_data.get("partial", [])),
        ("Missing Features", module_data.get("missing", [])),
        ("UI/UX Issues", module_data.get("ui_ux", [])),
        ("Suggested UI Improvements", module_data.get("suggestions", [])),
        ("Backend Dependency Notes", module_data.get("backend", [])),
    ]:
        ws.cell(row=r, column=1, value=section_title).font = SUBHEADER_FONT
        ws.cell(row=r, column=1).fill = SUBHEADER_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
        for item in items:
            sev = item if isinstance(item, tuple) else (item, "")
            text, severity = sev if len(sev) == 2 else (sev[0], "")
            ws.cell(row=r, column=1, value=text)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
            ws.cell(row=r, column=6, value=severity)
            r += 1
        r += 1

    set_col_widths(ws, [28, 18, 18, 18, 18, 14])
    return ws


def build_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Overview"

    ws["A1"] = "Finance Module — Frontend UI/UX Audit Report"
    ws["A1"].font = Font(bold=True, size=16, color="1A3A5C")
    ws.merge_cells("A1:F1")
    ws["A2"] = f"Audit Date: {date.today().isoformat()} | Scope: Frontend only (no backend assumptions)"
    ws.merge_cells("A2:F2")

    overview_sections = [
        ("Executive Summary", [
            "The Finance Operations module is a mature, self-contained React area under /finance/* with 10 routed pages, 56+ shared components, centralized API layer (financeAPI.js), RBAC permissions, and mock-data fallbacks.",
            "Core student payment workflows (dashboard, reports, verification, EMI, receipts, offline approval, gateway logs, communication, GST settings) are substantially implemented at the UI level.",
            "All pages currently operate on mock/in-memory data when VITE_FINANCE_USE_MOCK is enabled or in frontend-only mode. Real API contracts exist in financeAPI.js but backend integration is not production-ready.",
            "10 requested major features (refund management, scholarship/discount admin, faculty revenue, commission, incentives, accounting integration, etc.) are largely absent from the Finance module — some exist in adjacent modules (Bookstore, Sales Analytics).",
            "Navigation is complete for all 10 finance pages via sidebar (FINANCE_NAV_ITEMS) and FinanceLayout routes. No broken routes detected; minor workflow gaps (dead quick-action buttons, inconsistent page shells).",
        ]),
        ("Readiness Scores", []),
    ]

    r = 4
    for title, bullets in overview_sections:
        ws.cell(row=r, column=1, value=title).font = SECTION_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
        if title == "Readiness Scores":
            scores = [
                ("Overall Frontend Completion", "72%", "10 core pages built; major revenue features missing"),
                ("UI Readiness", "78%", "Consistent Figma design language; some layout inconsistencies"),
                ("UX Readiness", "68%", "Workflows exist but confirmation dialogs, breadcrumbs, mobile gaps"),
                ("Production Readiness", "42%", "Blocked on live APIs, real gateway, refund/accounting flows"),
                ("Final Readiness Score (weighted)", "58%", "Ready for backend integration phase, not production"),
            ]
            r = write_table(ws, r, ["Metric", "Score", "Notes", "", "", ""], scores) + 2
        else:
            for b in bullets:
                ws.cell(row=r, column=1, value=f"• {b}")
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
                r += 1
            r += 1

    r = write_table(
        ws,
        r,
        ["Module", "Completion %", "UI Status", "Data Mode", "Production Blocker", "Priority"],
        [
            ("1. Payment Dashboard", "78%", "Complete layout", "Mock + API fallback", "Live dashboard aggregation API", "High"),
            ("2. Student Payment Reports", "85%", "Feature-rich", "Mock + API fallback", "Reports CRUD + export API", "High"),
            ("3. Payment Verification Center", "88%", "Most complete workflow", "Mock + API fallback", "Verification queue API + file upload", "Critical"),
            ("4. EMI Management", "82%", "Strong edit modal", "Mock + API fallback", "EMI plan persistence API", "High"),
            ("5. Receipt Management", "80%", "Complete send flow UI", "Mock + API fallback", "Receipt generation + comms API", "High"),
            ("6. Student Finance Profiles", "65%", "Basic directory", "Mock only (notes mock)", "Profile CRUD API", "Medium"),
            ("7. Payment Attempt Logs", "72%", "Functional table", "Mock data", "Gateway webhook log API", "Medium"),
            ("8. Offline Payment Approval", "78%", "Approve/reject UI", "Mock + API fallback", "Offline approval workflow API", "High"),
            ("9. Payment Communication Logs", "70%", "Basic logs + send form", "Mock data", "Messaging service integration", "Medium"),
            ("10. GST & Invoice Settings", "75%", "Settings form", "Mock + API fallback", "GST config persistence API", "Medium"),
        ],
    )

    r = write_table(
        ws,
        r + 2,
        ["Recommended Next Steps (Frontend-Only, Pre-Backend)", "Owner", "Priority"],
        [
            ("Standardize all pages on FinancePageShell (Receipt Management uses PageBanner directly)", "Frontend", "High"),
            ("Wire batch filter on Payment Dashboard (batchFilter state exists but unused in API call)", "Frontend", "Medium"),
            ("Fix FinanceQuickActions 'Add Payment' — currently navigates to reports, not verification offline modal", "Frontend", "High"),
            ("Add reject confirmation dialog on Offline Payment Approval (approve has no confirm either)", "Frontend", "Medium"),
            ("Centralize PaymentStatusPill (Reports page) with FinanceStatusBadge", "Frontend", "Medium"),
            ("Add loading skeletons to EMI, Profiles, Attempt Logs, Communication pages", "Frontend", "Medium"),
            ("Define API contract document from financeAPI.js exports for backend team", "Frontend/Backend", "Critical"),
            ("Add breadcrumbs component across finance pages", "Frontend", "Low"),
        ],
    )

    set_col_widths(ws, [32, 14, 22, 22, 28, 12])

    # Module audit sheets
    modules = {
        "Dashboard Audit": {
            "name": "1. Payment Dashboard",
            "route": "/finance/dashboard",
            "component": "PaymentDashboardPage.jsx",
            "completion": "78%",
            "status": "Mostly Complete — Static/Mock driven",
            "data_source": "useFinanceDashboard → fetchPaymentDashboardByScope → financeMockData.buildFinanceDashboardPayload",
            "existing": [
                "FinancePageShell banner with refresh + export actions",
                "Center filter chips (All Centers + selected centers) via FinanceCenterFilterContext",
                "Course, month, batch dropdown filters (sticky filter bar)",
                "11 KPI stat cards (FinanceStatCard): revenue, collections, pending, overdue, verification, EMI, offline, success rate",
                "CenterPerformanceCards (overall view): best/worst center + conversion %",
                "Center revenue ranking list (top 5) with drill-down modal",
                "Charts: monthly revenue bar, payment status donut legend, course-wise revenue bar",
                "Recent successful + failed payment tables (PaginatedFigmaTable)",
                "FinanceActivityFeed widget",
                "FinanceQuickActions shortcut bar",
                "CenterDrillDownModal for center detail",
                "CSV export via exportFinanceCsv",
                "60-second auto-refresh polling",
            ],
            "implemented": [
                ("Center-scoped dashboard loading with compare/multi/overall scopes", "Completed"),
                ("KPI cards with INR formatting", "Mock data driven"),
                ("Interactive bar charts (click handlers)", "Static UI only"),
                ("Permission-gated export (canExport)", "Frontend integrated"),
                ("Loading skeleton (FinanceDashboardSkeleton)", "Completed"),
            ],
            "partial": [
                ("Batch filter dropdown rendered but NOT passed to useFinanceDashboard/API — filter has no effect", "High"),
                ("Donut chart is legend-only (no actual donut SVG/chart component — FinanceCharts.jsx exists but unused)", "Medium"),
                ("Center filter chips duplicate navbar CenterSelectorDropdown functionality", "Low"),
                ("Pending revenue fallback calculation: stats.pendingPayments * 5000 when pendingRevenue missing", "Medium"),
            ],
            "missing": [
                ("Real-time WebSocket updates for live collections", "Low"),
                ("Date range picker (only month dropdown from mock data)", "Medium"),
                ("Batch-wise revenue chart (batch filter non-functional)", "High"),
                ("Export all dashboard sections (currently exports recent payments only)", "Medium"),
                ("Mobile-optimized chart scrolling", "Medium"),
                ("Empty state when no center data", "Low"),
                ("Trend comparison (MoM/YoY)", "Medium"),
            ],
            "ui_ux": [
                ("11 stat cards in 4-column grid may overwhelm on tablet — no collapsible grouping", "Medium"),
                ("Batch filter misleads users — appears functional but does nothing", "High"),
                ("No breadcrumbs back to Finance home", "Low"),
                ("Custom inline BarChart/DonutLegend instead of shared FinanceCharts component", "Medium"),
                ("Sticky filter bar may overlap content on small screens", "Low"),
            ],
            "suggestions": [
                "Wire batchFilter into fetchPaymentDashboardByScope params",
                "Replace inline charts with FinanceCharts.jsx for consistency",
                "Add dashboard widget collapse/expand for KPI groups",
                "Add 'last updated' timestamp near refresh button",
                "Link failed payments rows to Payment Attempt Logs page",
            ],
            "backend": [
                "BLOCKED: Live /finance/payments/overall-dashboard, center/:id, compare-centers endpoints for scoped aggregation",
                "BLOCKED: Real-time collection totals and verification pending counts",
                "CAN DO FRONTEND: Wire batch filter, add date range UI, improve empty states, use FinanceCharts",
            ],
        },
        "Reports Audit": {
            "name": "2. Student Payment Reports",
            "route": "/finance/reports",
            "component": "StudentPaymentReportsPage.jsx",
            "completion": "85%",
            "status": "Feature Complete (Mock) — Best-in-module",
            "data_source": "fetchPaymentReports → mockReports (in-memory, mutated on edit)",
            "existing": [
                "Comprehensive filter panel: search, status, type, mode, course, center, date range, student ID, mobile",
                "16-column table with column visibility toggle",
                "Table density toggle (compact/comfortable)",
                "Row actions: View drawer, Edit modal, Generate receipt, Send reminder, Open profile",
                "PaymentViewDrawer with receipt download/print/resend/WhatsApp",
                "PaymentEditModal with status, amount adjustment, reason, comment",
                "FinanceExportToolbar CSV export",
                "Custom PaymentStatusPill + FinanceStatusBadge for verification/EMI",
                "Rich empty state with contextual messaging",
                "Zebra striping, sticky header, sticky last column, row animation",
                "RBAC: canEdit, canExport, canReceipts gates",
            ],
            "implemented": [
                ("Full CRUD-style status update via updatePaymentStatus", "Mock data driven"),
                ("Client-side filtering via filterPaymentReports", "Completed"),
                ("Debounced search (useDebouncedValue)", "Completed"),
                ("Cross-page profile drawer via FinanceOperationsContext", "Completed"),
                ("Receipt generate/resend flows", "Mock data driven"),
            ],
            "partial": [
                ("PaymentStatusPill duplicates FinanceStatusBadge with different styling (rounded-full vs rounded-md)", "Medium"),
                ("EditPaymentModal.jsx exists alongside PaymentEditModal.jsx — potential duplicate", "Low"),
                ("Refunded status in edit modal but not in filter dropdown", "Medium"),
                ("handlePrintReceipt opens /finance/receipts?preview=id — preview param not handled in ReceiptManagementPage", "High"),
            ],
            "missing": [
                ("Bulk actions (bulk reminder, bulk export selection)", "Medium"),
                ("Advanced filters: batch, verification status, EMI status", "Medium"),
                ("Pagination server-side (client-side only via PaginatedFigmaTable)", "Low"),
                ("Audit trail column / history of status changes", "Medium"),
                ("Refund initiation workflow", "High"),
                ("Invoice download (only receipt)", "Medium"),
            ],
            "ui_ux": [
                ("Filter panel spans 3 rows — high vertical space on mobile", "Medium"),
                ("5 icon-only action buttons per row — crowded on mobile, no action menu", "High"),
                ("Column visibility dropdown closes on outside click not implemented", "Low"),
                ("No confirmation before status change in edit modal", "Medium"),
            ],
            "suggestions": [
                "Consolidate status badges into FinanceStatusBadge",
                "Implement receipt preview route handler on ReceiptManagementPage",
                "Add row action overflow menu for mobile",
                "Add Refunded to status filter",
                "Add batch + verification status filters",
            ],
            "backend": [
                "BLOCKED: GET /finance/reports with server-side pagination/filtering",
                "BLOCKED: PATCH /finance/reports/:id/status with audit log",
                "BLOCKED: Receipt generation PDF service",
                "CAN DO FRONTEND: Fix preview route, unify badges, add confirmation dialogs, bulk UI",
            ],
        },
        "Verification Audit": {
            "name": "3. Payment Verification Center",
            "route": "/finance/verification",
            "component": "PaymentVerificationCenterPage.jsx",
            "completion": "88%",
            "status": "Most Complete Workflow",
            "data_source": "fetchVerificationQueue → financeVerificationData.INITIAL_VERIFICATION_QUEUE (in-memory)",
            "existing": [
                "Verification queue table: Payment ID, Student, Center, Course, Mode, Amount, UTR, Proof, Status, Actions",
                "FinanceFilterPanel: search + verification status filter + reset",
                "Actions per row: View proof, Approve, Reject, Escalate, Re-upload",
                "ProofViewerModal for payment proof images",
                "ConfirmActionDialog for approve confirmation",
                "VerificationRejectDialog with reason codes",
                "VerificationEscalateDialog with senior assignee",
                "AddOfflinePaymentModal — full offline payment + EMI setup (16 sub-components)",
                "VerificationStatusBadge with hover descriptions",
                "FinanceExportToolbar with VERIFICATION_QUEUE_EXPORT_COLUMNS",
                "Auto receipt generation on approve (when canReceipts)",
                "bumpRefresh to sync Student Payment Reports",
            ],
            "implemented": [
                ("Full approve/reject/escalate/reupload workflow", "Mock data driven"),
                ("Offline payment submission with EMI plan creation", "Mock data driven"),
                ("RBAC action gating (canApprove, canEdit)", "Completed"),
                ("Status-based action visibility (canActOnRow logic)", "Completed"),
            ],
            "partial": [
                ("ProofViewerModal shows placeholder when no real file URL — mock proof names only", "Medium"),
                ("AddOfflinePaymentModal title says 'Offline EMI Payment' even for full payment", "Low"),
                ("No table loading skeleton — only text 'Loading…'", "Medium"),
                ("Rejected rows show 'No further actions' — no reopen/review flow", "Low"),
            ],
            "missing": [
                ("Bulk approve/reject for batch verification", "Medium"),
                ("Assignment to specific verifier", "Medium"),
                ("SLA/deadline indicators for pending items", "Medium"),
                ("Filter by center, course, amount range, date", "Medium"),
                ("Audit log tab per verification record", "Medium"),
                ("Real file upload preview (PDF/image viewer)", "High"),
            ],
            "ui_ux": [
                ("5 action buttons per row wrap on narrow screens — needs action dropdown", "High"),
                ("Approve/Reject buttons lack double-confirm on reject (only approve has ConfirmActionDialog)", "Medium"),
                ("Table emptyMessage is plain text — not EmptyFinanceState component", "Low"),
            ],
            "suggestions": [
                "Add reject confirmation dialog matching approve pattern",
                "Add center/course/date filters to FinanceFilterPanel",
                "Use FinanceTableSkeleton during load",
                "Rename modal title dynamically based on payment type",
            ],
            "backend": [
                "BLOCKED: Verification queue API with file storage for payment proofs",
                "BLOCKED: POST /finance/verification/offline with document upload",
                "BLOCKED: Notification to student on approve/reject",
                "CAN DO FRONTEND: Reject confirm dialog, filters, skeleton, action menu",
            ],
        },
        "EMI Audit": {
            "name": "4. EMI Management",
            "route": "/finance/emi",
            "component": "EmiManagementPage.jsx + EmiEditModal.jsx",
            "completion": "82%",
            "status": "Substantially Complete — Mock driven",
            "data_source": "fetchEmiPlans / updateEmiPlan → MOCK_EMI_PLANS",
            "existing": [
                "Summary stat cards: active plans, total pending, due installments",
                "EMI plans table with progress bar, edit action",
                "All installments flat table (aggregated from all plans)",
                "EmiEditModal with: summary card, installment table, history drawer, early closure dialog",
                "Sub-components: EmiEditSummaryCard, EmiEditInstallmentTable, EmiEditHistoryDrawer, EmiEarlyClosureDialog",
                "Search filter (student/course/ID)",
                "FinanceExportToolbar with EMI_EXPORT_COLUMNS",
                "Collect installment + edit installment dialogs (in modal)",
            ],
            "implemented": [
                ("EMI plan list + installment view", "Mock data driven"),
                ("Edit modal with installment collect/closure", "Mock data driven"),
                ("Progress percentage visualization", "Completed"),
                ("RBAC canManageEmi gate on edit button", "Completed"),
            ],
            "partial": [
                ("No loading state on initial fetch", "Medium"),
                ("No filters beyond search (status, due date, overdue)", "Medium"),
                ("Installment table shows all plans — can be very long without plan-level filter", "Medium"),
                ("No link from overdue installment to send reminder", "Low"),
            ],
            "missing": [
                ("Create new EMI plan from this page (only via Verification offline flow)", "Medium"),
                ("EMI reminder scheduling UI", "Medium"),
                ("Overdue escalation workflow", "Medium"),
                ("EMI analytics charts (collection trend)", "Low"),
                ("Export installments separately", "Low"),
            ],
            "ui_ux": [
                ("Two large tables stacked — no tabs to switch Plans vs Installments", "Medium"),
                ("Edit button is icon-only without label", "Low"),
                ("No empty state component when no plans", "Low"),
            ],
            "suggestions": [
                "Add tabs: Plans | Installments | Overdue",
                "Add status/due date filters",
                "Add loading skeleton",
                "Link overdue rows to Communication Logs reminder form",
            ],
            "backend": [
                "BLOCKED: GET/PUT /finance/emi with installment mutation persistence",
                "BLOCKED: EMI reminder trigger API",
                "CAN DO FRONTEND: Tabs, filters, skeletons, reminder deep-link",
            ],
        },
        "Receipt Audit": {
            "name": "5. Receipt Management",
            "route": "/finance/reports",
            "component": "ReceiptManagementPage.jsx + receipt-center/*",
            "completion": "80%",
            "status": "Complete UI — Layout inconsistency",
            "data_source": "fetchCompletedReceipts → filterCompletedReceipts(mockReports)",
            "existing": [
                "PageBanner (NOT FinancePageShell) with receipt count badge",
                "Filter panel: search, course, payment type, center, date range",
                "ReceiptCenterTable with send actions",
                "SendReceiptDialog: Email/WhatsApp/SMS channel selection",
                "CommunicationStatusChips per channel status",
                "CompletionReceiptDocument for printable receipt",
                "ReceiptPreview.jsx modal component",
                "FinanceExportToolbar",
            ],
            "implemented": [
                ("Completed receipts list (fully paid + EMI completed only)", "Mock data driven"),
                ("Send receipt communication flow", "Mock data driven"),
                ("Client-side filtering via filterReceiptCenterRows", "Completed"),
            ],
            "partial": [
                ("Uses PageBanner directly instead of FinancePageShell — inconsistent padding/layout", "Medium"),
                ("?preview= query param from Reports print action NOT implemented", "High"),
                ("No loading state on fetch", "Medium"),
                ("ReceiptPreview component exists but not wired to page route", "High"),
            ],
            "missing": [
                ("Bulk send receipts", "Medium"),
                ("Receipt template customization preview", "Medium"),
                ("Download PDF batch export", "Medium"),
                ("Receipt re-generation for corrections", "Medium"),
                ("GST breakdown on receipt document", "Low"),
            ],
            "ui_ux": [
                ("Different page shell breaks visual continuity with other finance pages", "Medium"),
                ("Filter labels use uppercase xs text — differs from other pages", "Low"),
                ("No empty state when no completed receipts", "Medium"),
            ],
            "suggestions": [
                "Migrate to FinancePageShell for consistency",
                "Handle ?preview=id query param to open ReceiptPreview",
                "Add EmptyFinanceState",
                "Add loading skeleton",
            ],
            "backend": [
                "BLOCKED: Receipt PDF generation service",
                "BLOCKED: Multi-channel delivery API (Email/WhatsApp/SMS)",
                "CAN DO FRONTEND: Shell migration, preview route, empty/loading states",
            ],
        },
        "Finance Profiles": {
            "name": "6. Student Finance Profiles",
            "route": "/finance/profiles",
            "component": "StudentFinanceProfilesPage.jsx + StudentFinanceProfilePanel.jsx",
            "completion": "65%",
            "status": "Basic — Partially mock",
            "data_source": "fetchStudentFinanceProfiles → MOCK_STUDENT_PROFILES",
            "existing": [
                "Student profiles table: name, ID, mobile, branch, total paid, pending",
                "Course finance history table (flattened from all profiles)",
                "FinanceSlideDrawer for profile detail + admin note textarea",
                "Global StudentFinanceProfilePanel slide-over (via FinanceOperationsContext)",
                "Search filter",
            ],
            "implemented": [
                ("Profile directory listing", "Mock data driven"),
                ("Course history aggregation", "Completed"),
                ("Cross-page profile panel from Reports/Offline pages", "Mock data driven"),
            ],
            "partial": [
                ("handleSaveNote shows toast 'Profile note saved (mock)' — no API call", "High"),
                ("Edit pencil opens drawer but limited edit capabilities (notes only)", "Medium"),
                ("No export functionality", "Medium"),
                ("StudentFinanceProfilePanel vs FinanceSlideDrawer — duplicate profile UIs", "Medium"),
            ],
            "missing": [
                ("Wallet balance integration", "Medium"),
                ("Scholarship/discount history per student", "High"),
                ("Payment timeline visualization", "Medium"),
                ("Link to EMI plan from profile", "Medium"),
                ("Filters: branch, payment status, course", "Medium"),
                ("Profile export", "Low"),
            ],
            "ui_ux": [
                ("Two profile UIs (page drawer + global panel) may confuse users", "Medium"),
                ("No loading or empty states", "Medium"),
                ("Tables lack row click to open profile (only pencil icon)", "Low"),
            ],
            "suggestions": [
                "Unify profile UI into single StudentFinanceProfilePanel",
                "Wire save note to API placeholder with optimistic UI",
                "Add filters and export",
                "Add payment timeline tab in profile drawer",
            ],
            "backend": [
                "BLOCKED: GET /finance/profiles with note persistence",
                "BLOCKED: Student wallet/scholarship integration",
                "CAN DO FRONTEND: Unified profile panel, filters, timeline UI (mock), loading states",
            ],
        },
        "Attempt Logs": {
            "name": "7. Payment Attempt Logs (Gateway Logs)",
            "route": "/finance/attempts",
            "component": "PaymentAttemptLogsPage.jsx",
            "completion": "72%",
            "status": "Functional table — Mock gateway data",
            "data_source": "fetchPaymentAttemptLogs → MOCK_ATTEMPT_LOGS (buildAttempts in financeMockData)",
            "existing": [
                "Table columns: Student, Course, Attempt #, Txn ID, Mode, Amount, Gateway status, Gateway message, Status, Date",
                "Search filter (student, txn)",
                "Status filter (Success/Failed)",
                "Mode filter (UPI/Card only)",
                "FinanceExportToolbar CSV export",
            ],
            "implemented": [
                ("Gateway attempt log listing", "Mock data driven"),
                ("Client-side filter + search", "Completed"),
                ("FinanceStatusBadge for status column", "Completed"),
            ],
            "partial": [
                ("No loading skeleton or loading indicator", "Medium"),
                ("Mode filter limited to UPI/Card — mock has more modes", "Low"),
                ("No detail drawer for full gateway response JSON", "Medium"),
                ("gatewayResponse field in mock data not displayed in UI", "Medium"),
            ],
            "missing": [
                ("Gateway provider filter (Razorpay, Cashfree, etc.)", "High"),
                ("Retry payment action", "Medium"),
                ("Date range filter", "Medium"),
                ("Center/course filters", "Medium"),
                ("Failed attempt analytics chart", "Low"),
                ("Link to student report from log row", "Medium"),
            ],
            "ui_ux": [
                ("Wide table may overflow horizontally on mobile — no sticky columns", "Medium"),
                ("No empty state component", "Low"),
                ("Gateway message column may truncate long error strings", "Low"),
            ],
            "suggestions": [
                "Add gateway response detail modal",
                "Add date range + gateway provider filters",
                "Add FinanceLinkedActions to jump to student profile/report",
                "Add failed attempts summary cards at top",
            ],
            "backend": [
                "BLOCKED: GET /finance/attempts from payment gateway webhook store",
                "BLOCKED: Gateway retry API",
                "CAN DO FRONTEND: Detail modal, filters, summary cards, linked actions",
            ],
        },
        "Offline Approval": {
            "name": "8. Offline Payment Approval",
            "route": "/finance/offline-approval",
            "component": "OfflinePaymentApprovalPage.jsx",
            "completion": "78%",
            "status": "Workflow UI complete — Mock driven",
            "data_source": "fetchOfflineApprovals / approveOfflinePayment → MOCK_OFFLINE_REQUESTS",
            "existing": [
                "3 summary stat cards: pending, total, approved",
                "FinanceFilterPanel: search + status filter",
                "Table: Request ID, Student, Center, Course, Amount, Mode, UTR, Proof, Status, Decision",
                "ProofViewerModal for proof preview",
                "Approve/Reject inline buttons",
                "FinanceLinkedActions for cross-navigation",
                "FinanceExportToolbar",
                "Default filter: Pending Approval",
            ],
            "implemented": [
                ("Approve/reject decision flow", "Mock data driven"),
                ("Proof preview modal", "Completed"),
                ("RBAC canApprove gate", "Completed"),
                ("bumpRefresh on decision", "Completed"),
            ],
            "partial": [
                ("No confirmation dialog before approve/reject — immediate action", "High"),
                ("Reject lacks reason capture (unlike VerificationRejectDialog)", "High"),
                ("Overlap with Payment Verification Center — two offline workflows may confuse users", "Medium"),
                ("No loading state", "Medium"),
            ],
            "missing": [
                ("Escalation workflow (exists in Verification but not here)", "Medium"),
                ("Bulk approve", "Low"),
                ("Add offline payment entry (only in Verification page)", "Medium"),
                ("Audit trail per request", "Medium"),
            ],
            "ui_ux": [
                ("Approve/Reject buttons stacked with FinanceLinkedActions — cramped action column", "Medium"),
                ("No empty state for zero pending requests", "Low"),
                ("Relationship to Verification Center not explained in UI", "High"),
            ],
            "suggestions": [
                "Add ConfirmActionDialog + reject reason dialog (reuse VerificationRejectDialog)",
                "Add info banner explaining difference vs Verification Center",
                "Add 'Add Offline Payment' shortcut linking to Verification page",
            ],
            "backend": [
                "BLOCKED: GET/PATCH /finance/offline-approvals",
                "BLOCKED: Clarify backend model: offline approval vs verification queue",
                "CAN DO FRONTEND: Confirm dialogs, reject reasons, info banner, loading states",
            ],
        },
        "Communication Logs": {
            "name": "9. Payment Communication Logs",
            "route": "/finance/communication",
            "component": "PaymentCommunicationLogsPage.jsx",
            "completion": "70%",
            "status": "Basic — Mock driven",
            "data_source": "fetchCommunicationLogs / sendPaymentReminder → MOCK_COMMUNICATION_LOGS",
            "existing": [
                "Send payment reminder form: mobile, email, channel (WhatsApp/Email/SMS), send button",
                "Communication logs table: Log ID, Type, Recipient, Channel, Status, Sent timestamp",
                "Search + channel filter (Email/WhatsApp — SMS missing from filter)",
                "FinanceExportToolbar",
            ],
            "implemented": [
                ("Reminder send form + log listing", "Mock data driven"),
                ("Client-side filter", "Completed"),
            ],
            "partial": [
                ("Reminder form lacks validation (empty mobile/email allowed)", "Medium"),
                ("No template preview for reminder messages", "Medium"),
                ("SMS in send form but not in channel filter", "Low"),
                ("No loading state", "Medium"),
            ],
            "missing": [
                ("Message template management", "High"),
                ("Delivery status tracking (delivered/read/failed)", "Medium"),
                ("Scheduled reminders", "Medium"),
                ("Bulk reminder from selected students", "Medium"),
                ("Link log entry to student/payment record", "Medium"),
                ("Receipt communication logs merged view", "Low"),
            ],
            "ui_ux": [
                ("Send form and logs on same page — no visual separation of compose vs history", "Low"),
                ("No confirmation before sending reminder", "Medium"),
                ("No empty state", "Low"),
            ],
            "suggestions": [
                "Add form validation (require mobile OR email)",
                "Add SMS to channel filter",
                "Add template selector dropdown",
                "Add delivery status badges",
            ],
            "backend": [
                "BLOCKED: Messaging service API (WhatsApp/Email/SMS providers)",
                "BLOCKED: Template management API",
                "CAN DO FRONTEND: Validation, template UI mock, status badges, confirm dialog",
            ],
        },
        "GST Settings": {
            "name": "10. GST & Invoice Settings",
            "route": "/finance/gst-settings",
            "component": "GstInvoiceSettingsPage.jsx",
            "completion": "75%",
            "status": "Settings form — Mock persistence",
            "data_source": "fetchGstSettings / updateGstSettings → MOCK_GST_SETTINGS",
            "existing": [
                "Global tax settings: GST %, invoice prefix, receipt prefix, tax enabled checkbox",
                "Branch GST section: per-branch GST enabled toggle + GSTIN input",
                "Save button in page banner (form submit)",
                "RBAC canManageGst disables inputs for unauthorized roles",
            ],
            "implemented": [
                ("Settings load and save flow", "Mock data driven"),
                ("Per-branch GST configuration UI", "Completed"),
                ("Permission-gated editing", "Completed"),
            ],
            "partial": [
                ("No GSTIN format validation (15-char pattern)", "Medium"),
                ("No invoice preview/sample", "Medium"),
                ("No HSN/SAC code configuration", "Low"),
                ("No audit log of settings changes", "Low"),
            ],
            "missing": [
                ("Invoice template editor", "High"),
                ("Tax exemption rules per course/category", "Medium"),
                ("Multi-GST rate support (5%, 12%, 18%)", "Medium"),
                ("Integration with receipt generation preview", "Medium"),
                ("Export settings backup", "Low"),
            ],
            "ui_ux": [
                ("Save button uses bg-[#1a3a5c] — differs from primary #246392 on other pages", "Low"),
                ("No unsaved changes warning on navigate away", "Medium"),
                ("No success inline feedback beyond toast", "Low"),
            ],
            "suggestions": [
                "Add GSTIN validation regex",
                "Add invoice preview panel",
                "Add unsaved changes guard",
                "Standardize save button color to #246392",
            ],
            "backend": [
                "BLOCKED: GET/PUT /finance/gst-settings persistence",
                "BLOCKED: Invoice template rendering service",
                "CAN DO FRONTEND: Validation, preview mock, unsaved guard",
            ],
        },
    }

    # Fix Receipt route typo
    modules["Receipt Audit"]["route"] = "/finance/receipts"

    for sheet_name, data in modules.items():
        add_module_sheet(wb, sheet_name, data)

    # Major Features sheet
    ws_mf = wb.create_sheet("Major Features")
    r = write_table(
        ws_mf,
        1,
        ["Feature", "Location in Codebase", "Completion %", "Status", "Severity Gap", "Notes"],
        [
            ("Refund Management", "Finance: status option only (PaymentEditModal). Bookstore: BookstorePaymentsPage.jsx has refund modal", "18%", "Not in Finance module", "Critical", "No dedicated finance refund workflow, approval, or audit trail"),
            ("Scholarship & Discount Management", "studentCourseFeeProfiles.js + OfflinePaymentStudentSummary.jsx (display only)", "25%", "Partial — display in offline payment", "High", "No admin CRUD for scholarships/discounts; hardcoded per course in mock data"),
            ("Book Purchase Billing", "Separate module: /admin/bookstore/payments (BookstorePaymentsPage.jsx)", "40%", "Separate module, not integrated", "High", "Bookstore payments isolated from finance dashboard/reports"),
            ("Referral and Commission Tracking", "Sales Analytics: salesAnalyticsMockData.js (referral source). Not in finance.", "5%", "Missing from Finance", "Critical", "No commission calculation, payout tracking, or referral finance reports"),
            ("Faculty Revenue Sharing", "Not found in finance module", "0%", "Missing", "Critical", "Faculty referenced in batch/test modules only; no revenue share UI"),
            ("Batch-wise Revenue Analytics", "FINANCE_BATCHES constant + dashboard batchFilter (non-functional)", "20%", "Partial — filter UI only", "High", "Batch dropdown exists but not wired to data/API"),
            ("Sales Team Incentive Tracking", "Not found", "0%", "Missing", "Critical", "No incentive rules, targets, or payout UI in finance or sales modules"),
            ("Accounting Software Integration", "moduleRoutes: system/api-integrations (generic placeholder)", "5%", "Placeholder only", "High", "No Tally/Zoho/QuickBooks export or sync UI in finance"),
            ("Payment Gateway Logs", "PaymentAttemptLogsPage.jsx (/finance/attempts)", "72%", "Implemented in Finance", "Medium", "Functional mock UI; needs live gateway webhook data"),
            ("Free-to-Paid User Conversion Tracking", "Dashboard: conversionPct in CenterPerformanceCards. Sales: ConversionFunnelPage", "30%", "Split across modules", "Medium", "No unified free-to-paid funnel in finance; center conversion % only on dashboard"),
        ],
    )
    set_col_widths(ws_mf, [28, 36, 12, 18, 12, 40])

    # UI UX Issues consolidated
    ws_ux = wb.create_sheet("UI UX Issues")
    ux_issues = [
        ("Navigation", "FinanceQuickActions 'Add Payment' navigates to reports instead of offline payment entry", "High", "PaymentDashboardPage / FinanceQuickActions.jsx"),
        ("Navigation", "No breadcrumbs on any finance page", "Low", "All finance pages"),
        ("Navigation", "Offline Approval vs Verification Center relationship unclear to users", "High", "OfflinePaymentApprovalPage / PaymentVerificationCenterPage"),
        ("Layout", "ReceiptManagementPage uses PageBanner instead of FinancePageShell", "Medium", "ReceiptManagementPage.jsx"),
        ("Layout", "Dashboard batch filter appears functional but has no effect", "High", "PaymentDashboardPage.jsx"),
        ("Layout", "11 KPI cards on dashboard may overwhelm mobile users", "Medium", "PaymentDashboardPage.jsx"),
        ("Components", "PaymentStatusPill duplicates FinanceStatusBadge with different styles", "Medium", "StudentPaymentReportsPage.jsx"),
        ("Components", "EditPaymentModal.jsx and PaymentEditModal.jsx both exist", "Low", "components/finance/"),
        ("Components", "FinanceCharts.jsx exists but dashboard uses inline BarChart", "Medium", "FinanceCharts.jsx / PaymentDashboardPage.jsx"),
        ("Components", "Two student profile UIs: FinanceSlideDrawer + StudentFinanceProfilePanel", "Medium", "StudentFinanceProfilesPage / FinanceOperationsContext"),
        ("Tables", "Reports page 5 icon actions per row — crowded on mobile", "High", "StudentPaymentReportsPage.jsx"),
        ("Tables", "Verification queue 5 action buttons wrap awkwardly", "High", "PaymentVerificationCenterPage.jsx"),
        ("Tables", "Payment Attempt Logs wide table lacks horizontal scroll hints", "Medium", "PaymentAttemptLogsPage.jsx"),
        ("Forms", "Communication reminder form lacks validation", "Medium", "PaymentCommunicationLogsPage.jsx"),
        ("Forms", "GST settings save button color inconsistent (#1a3a5c vs #246392)", "Low", "GstInvoiceSettingsPage.jsx"),
        ("Dialogs", "Offline approval approve/reject lacks confirmation dialog", "High", "OfflinePaymentApprovalPage.jsx"),
        ("Dialogs", "Reject on offline approval lacks reason capture", "High", "OfflinePaymentApprovalPage.jsx"),
        ("Dialogs", "Receipt preview ?preview= query not handled", "High", "ReceiptManagementPage.jsx"),
        ("States", "Missing loading skeletons on EMI, Profiles, Attempts, Communication, Receipt pages", "Medium", "Multiple pages"),
        ("States", "Missing empty states on several pages (EMI, Attempts, Communication, Offline)", "Medium", "Multiple pages"),
        ("Accessibility", "Icon-only action buttons rely on title attr — no aria-label on some", "Medium", "ReportActionButton has aria-label; verification actions do not"),
        ("Responsive", "Filter panels on Reports page consume excessive vertical space on mobile", "Medium", "StudentPaymentReportsPage.jsx"),
        ("Color", "Status badge shapes inconsistent: rounded-full vs rounded-md across pages", "Low", "FinanceStatusBadge vs PaymentStatusPill"),
    ]
    write_table(ws_ux, 1, ["Category", "Issue", "Severity", "Component/File"], ux_issues)
    set_col_widths(ws_ux, [14, 52, 10, 36])

    # Recommendations
    ws_rec = wb.create_sheet("Recommendations")
    recs = [
        ("High", "Frontend", "Define and publish API contract document from financeAPI.js for backend team", "Unblocks parallel backend development"),
        ("High", "Frontend", "Fix dead/misleading UI: batch filter, Add Payment quick action, receipt preview route", "Prevents user confusion during UAT"),
        ("High", "Frontend", "Add confirmation + reason dialogs to Offline Payment Approval", "Matches Verification Center UX quality"),
        ("High", "Frontend+Design", "Plan Refund Management page within finance module (reference BookstorePaymentsPage pattern)", "Critical missing feature"),
        ("High", "Architecture", "Standardize ReceiptManagementPage on FinancePageShell", "Design consistency"),
        ("Medium", "Frontend", "Create shared FinanceActionMenu component for row actions (mobile-friendly overflow)", "Reusable across Reports, Verification, Offline"),
        ("Medium", "Frontend", "Consolidate FinanceStatusBadge as single status chip — remove PaymentStatusPill", "Design consistency"),
        ("Medium", "Frontend", "Add FinanceTableSkeleton to all data pages lacking loading states", "UX polish"),
        ("Medium", "Frontend", "Add FinanceBreadcrumbs component to FinancePageShell", "Navigation clarity"),
        ("Medium", "Architecture", "Unify StudentFinanceProfilePanel and FinanceSlideDrawer profile views", "Reduce duplicate code"),
        ("Medium", "Product", "Clarify Offline Approval vs Verification Center — merge or add explanatory banners", "Workflow clarity"),
        ("Low", "Frontend", "Wire FinanceCharts.jsx on dashboard instead of inline chart components", "Maintainability"),
        ("Low", "Frontend", "Remove or merge duplicate EditPaymentModal.jsx if unused", "Code hygiene"),
        ("Low", "Frontend", "Add unsaved changes guard on GST settings form", "Prevent data loss"),
    ]
    write_table(ws_rec, 1, ["Priority", "Area", "Recommendation", "Rationale"], recs)
    set_col_widths(ws_rec, [10, 16, 48, 32])

    # Backend Dependencies
    ws_bd = wb.create_sheet("Backend Dependencies")
    deps = [
        ("Payment Dashboard", "GET /finance/payments/overall-dashboard, /center/:id, /compare-centers", "Critical", "Not Ready", "Frontend can wire batch filter UI; aggregation logic exists in mock"),
        ("Student Payment Reports", "GET /finance/reports, PATCH /finance/reports/:id/status", "Critical", "Not Ready", "Full table UI ready; needs pagination contract"),
        ("Verification Center", "GET/POST /finance/verification/*, file upload for proofs", "Critical", "Not Ready", "Full workflow UI ready"),
        ("EMI Management", "GET/PUT /finance/emi, POST create plan", "High", "Not Ready", "EmiEditModal ready for API"),
        ("Receipt Management", "GET /finance/receipts/completed, POST send, PDF generation", "High", "Not Ready", "Send dialog + document template ready"),
        ("Finance Profiles", "GET /finance/profiles, PATCH notes", "Medium", "Not Ready", "Note save is mock-only"),
        ("Attempt Logs", "GET /finance/attempts from gateway webhooks", "Medium", "Not Ready", "Table UI ready"),
        ("Offline Approval", "GET/PATCH /finance/offline-approvals", "High", "Not Ready", "Needs backend model clarification vs verification"),
        ("Communication Logs", "GET/POST communication + messaging provider", "Medium", "Not Ready", "Form UI ready; needs template API"),
        ("GST Settings", "GET/PUT /finance/gst-settings", "Medium", "Not Ready", "Form ready"),
        ("Refund Management", "POST /finance/refunds, GET refund queue", "Critical", "Missing", "No finance UI — only bookstore stub"),
        ("Scholarship/Discount", "CRUD /finance/scholarships, /finance/discounts", "High", "Missing", "Display-only in offline payment"),
        ("Accounting Integration", "Export/sync API (Tally, Zoho)", "High", "Missing", "No finance UI"),
        ("Commission/Referral", "GET /finance/commissions, referral payouts", "Critical", "Missing", "Exists in sales analytics only"),
        ("Faculty Revenue Share", "GET /finance/faculty-revenue", "Critical", "Missing", "No UI"),
        ("Incentive Tracking", "GET /finance/incentives", "Critical", "Missing", "No UI"),
    ]
    write_table(ws_bd, 1, ["Module/Feature", "API Endpoints Needed", "Priority", "Backend Status", "Frontend Can Proceed"], deps)
    set_col_widths(ws_bd, [22, 42, 10, 12, 36])

    # Navigation Flow Audit
    ws_nav = wb.create_sheet("Navigation Flow")
    nav_rows = [
        ("/finance → /finance/dashboard", "OK", "Redirect in moduleRoutes.jsx + FinanceLayout index", "None"),
        ("/finance/dashboard", "OK", "PaymentDashboardPage — in sidebar", "None"),
        ("/finance/reports", "OK", "StudentPaymentReportsPage — in sidebar", "None"),
        ("/finance/verification", "OK", "PaymentVerificationCenterPage — in sidebar", "None"),
        ("/finance/emi", "OK", "EmiManagementPage — in sidebar", "None"),
        ("/finance/receipts", "OK", "ReceiptManagementPage — in sidebar", "None"),
        ("/finance/profiles", "OK", "StudentFinanceProfilesPage — in sidebar", "None"),
        ("/finance/attempts", "OK", "PaymentAttemptLogsPage — in sidebar", "None"),
        ("/finance/offline-approval", "OK", "OfflinePaymentApprovalPage — in sidebar", "None"),
        ("/finance/communication", "OK", "PaymentCommunicationLogsPage — in sidebar", "None"),
        ("/finance/gst-settings", "OK", "GstInvoiceSettingsPage — in sidebar", "None"),
        ("/finance/* unknown", "OK", "NestedRouteRedirect → dashboard", "None"),
        ("Quick Action: Add Payment", "BROKEN", "FinanceQuickActions → goToFinance('reports')", "Should open Verification offline modal or navigate to verification"),
        ("Quick Action: Send Reminder", "OK", "→ /finance/communication", "None"),
        ("Reports: Print Receipt", "BROKEN", "window.open /finance/receipts?preview=id", "ReceiptManagementPage ignores preview param"),
        ("FinanceOperationsContext goToFinance", "OK", "All FINANCE_ROUTES keys valid", "None"),
        ("Global center filter navbar", "OK", "FinanceCenterFilterContext in App.jsx", "Duplicates dashboard center chips"),
        ("RBAC sidebar visibility", "OK", "navigation.js finance group + rbacAccess.js", "Some roles may lack finance nav access — verify QA"),
        ("Cross-module: Bookstore payments", "DISCONNECTED", "/admin/bookstore/payments separate module", "Not linked from finance nav"),
        ("Cross-module: Sales conversion", "DISCONNECTED", "/sales-analytics/funnel separate", "Not linked from finance dashboard"),
    ]
    write_table(ws_nav, 1, ["Route/Flow", "Status", "Implementation", "Issue"], nav_rows)
    set_col_widths(ws_nav, [32, 10, 36, 32])

    # Reusable Component Audit
    ws_reuse = wb.create_sheet("Reusable Components")
    reuse_rows = [
        ("Tables", "PaginatedFigmaTable", "Used across all finance pages", "Good — centralized", "Add shared FinanceDataTable wrapper with standard empty/loading"),
        ("Filters", "FinanceFilterPanel", "Verification, Offline Approval", "Partial", "Extract Reports filter grid into FinanceFilterPanel or FinanceAdvancedFilters"),
        ("Filters", "Inline filter grids", "Reports, Receipt, Attempt Logs, Communication", "Duplicated", "Centralize filter row component"),
        ("Status chips", "FinanceStatusBadge", "Most pages", "Good", "Remove PaymentStatusPill duplicate on Reports"),
        ("Status chips", "VerificationStatusBadge", "Verification page", "Good", "Keep separate — domain-specific"),
        ("Dialogs", "ConfirmActionDialog", "Verification approve", "Partial", "Reuse on Offline Approval approve/reject"),
        ("Dialogs", "VerificationRejectDialog", "Verification", "Good", "Reuse on Offline reject"),
        ("Dialogs", "ProofViewerModal", "Verification, Offline", "Good", "Centralized"),
        ("Export", "FinanceExportToolbar", "Most pages", "Good", "Add to Profiles, EMI installment export"),
        ("Search bars", "Inline inputs", "All pages", "Duplicated", "Create FinanceSearchInput component"),
        ("Analytics cards", "FinanceStatCard", "Dashboard, EMI, Offline", "Good", "Use on Attempt Logs (failed summary)"),
        ("Graph containers", "FinanceCharts.jsx", "Unused", "Gap", "Wire on dashboard; add to EMI analytics"),
        ("Empty states", "EmptyFinanceState.jsx", "Reports only (custom inline)", "Underused", "Apply to all table pages"),
        ("Pagination", "PaginatedFigmaTable", "All tables", "Good", "None"),
        ("Page shell", "FinancePageShell", "9/10 pages", "Gap", "Migrate ReceiptManagementPage"),
        ("Drawers", "FinanceSlideDrawer + PaymentViewDrawer", "Profiles, Reports", "Partial overlap", "Standardize drawer API"),
        ("Profile panel", "StudentFinanceProfilePanel", "Global via context", "Good", "Replace FinanceSlideDrawer on Profiles page"),
        ("Quick actions", "FinanceQuickActions + FinanceLinkedActions", "Dashboard, Offline", "Good", "Fix Add Payment target"),
        ("Loading", "FinanceTableSkeleton / FinanceDashboardSkeleton", "Dashboard, Reports", "Partial", "Apply to all data pages"),
        ("Column picker", "ColumnVisibilityToggle vs inline", "Reports inline only", "Duplicated", "Use ColumnVisibilityToggle component"),
    ]
    write_table(ws_reuse, 1, ["Category", "Component", "Current Usage", "Assessment", "Recommendation"], reuse_rows)
    set_col_widths(ws_reuse, [14, 28, 28, 14, 36])

    # Architecture Review
    ws_arch = wb.create_sheet("Architecture Review")
    arch_rows = [
        ("Folder structure", "Good", "src/pages/finance (10 pages), src/components/finance (56 files), subfolders emi-edit/, offline-payment/, receipt-center/", "Clear domain separation"),
        ("API layer", "Good", "Single financeAPI.js with tryApi mock fallback pattern", "Ready for backend swap via VITE_FINANCE_USE_MOCK"),
        ("State management", "Adequate", "React Context: FinanceOperationsContext, FinanceCenterFilterContext. Local useState per page.", "No Redux — acceptable for module size; consider React Query for cache/invalidation"),
        ("Hooks", "Good", "useFinanceDashboard, useFinancePermissions, useOfflinePaymentEmiForm", "Extend with useFinanceTableLoader pattern"),
        ("Constants", "Good", "financeNav, financeConstants, financePermissions, financeVerification", "Well organized"),
        ("Utils", "Good", "financeFilters, financeExport, financeRecordModel, emiSchedule, receiptCompletion", "Some reused in bookstore — good"),
        ("Naming", "Issue", "EditPaymentModal.jsx AND PaymentEditModal.jsx coexist", "Remove unused duplicate"),
        ("Hardcoded values", "Issue", "Admin names ('Admin', 'Verifier') hardcoded in API calls", "Use auth context user name"),
        ("Duplicate code", "Issue", "Filter grids, status pills, profile drawers duplicated", "See Reusable Components sheet"),
        ("Scalability", "Concern", "In-memory mock state in financeAPI.js won't scale for testing large datasets", "Add MSW or seed volume controls for QA"),
        ("RBAC", "Good", "financePermissions.js + useFinancePermissions hook", "Verify all action buttons respect permissions"),
        ("Error handling", "Adequate", "FinanceErrorBoundary wraps routes; toast on API errors", "Add retry UI on failed loads"),
        ("Documentation", "Good", "docs/finance/Finance-Operations-Flow-Documentation.md + QA CSVs", "Keep synced with UI changes"),
    ]
    write_table(ws_arch, 1, ["Area", "Rating", "Details", "Notes"], arch_rows)
    set_col_widths(ws_arch, [18, 10, 48, 32])

    # Design Consistency
    ws_design = wb.create_sheet("Design Consistency")
    design_rows = [
        ("Typography", "Mostly consistent", "text-sm body, text-xs labels, font-bold headings in #246392", "Receipt page uses text-xs font-semibold #555 for labels — slight deviation"),
        ("Colors — Primary", "Consistent", "#246392 primary, #55ace7 accent, #1a3a5c dark", "GST save button uses #1a3a5c — acceptable variant"),
        ("Colors — Status", "Mostly consistent", "Green #69df66 paid, amber #efb36d partial, red #df8284 failed", "PaymentStatusPill on Reports uses same hex but different border-radius"),
        ("Button variants", "Inconsistent", "Banner actions: white/20 ring. Page buttons: rounded-lg #246392. Verification: inline colored xs buttons", "Standardize primary/secondary/destructive button components"),
        ("Dialog styles", "Mixed", "Modal (ui/Modal) for offline payment; ConfirmActionDialog from batch-management; custom dialogs for verification", "Create finance-specific dialog wrappers"),
        ("Input fields", "Consistent", "rounded-lg border-slate-200 px-3 py-2 text-sm", "Receipt page adds h-10 + focus ring — slightly richer"),
        ("Table styling", "Consistent", "PaginatedFigmaTable with figma styling", "Reports adds zebra/sticky — consider as default props"),
        ("Icons", "Consistent", "Lucide icons throughout", "None"),
        ("Spacing", "Consistent", "gap-4 p-4 sm:gap-5 sm:p-6 via FinancePageShell", "Receipt page manual padding — migrate to shell"),
        ("Page banners", "Inconsistent", "FinancePageShell → PageBanner for 9 pages; Receipt uses PageBanner directly", "Standardize"),
        ("Card shadows", "Consistent", "shadow-[0_8px_24px_rgba(15,23,42,0.08)]", "None"),
        ("Status badge shape", "Inconsistent", "rounded-md (FinanceStatusBadge) vs rounded-full (PaymentStatusPill, Reports verification col)", "Unify to one shape"),
    ]
    write_table(ws_design, 1, ["Element", "Consistency", "Observation", "Fix"], design_rows)
    set_col_widths(ws_design, [18, 14, 42, 32])

    # Priority Fixes sheets
    ws_hi = wb.create_sheet("High Priority Fixes")
    hi = [r for r in recs if r[0] == "High"]
    write_table(ws_hi, 1, ["Priority", "Area", "Fix", "Rationale"], hi)
    set_col_widths(ws_hi, [10, 16, 48, 32])

    ws_med = wb.create_sheet("Medium Priority Fixes")
    med = [r for r in recs if r[0] == "Medium"]
    write_table(ws_med, 1, ["Priority", "Area", "Fix", "Rationale"], med)
    set_col_widths(ws_med, [10, 16, 48, 32])

    ws_lo = wb.create_sheet("Low Priority Fixes")
    lo = [r for r in recs if r[0] == "Low"]
    write_table(ws_lo, 1, ["Priority", "Area", "Fix", "Rationale"], lo)
    set_col_widths(ws_lo, [10, 16, 48, 32])

    return wb


if __name__ == "__main__":
    import os
    out_dir = os.path.dirname(os.path.abspath(__file__))
    out_path = os.path.join(out_dir, OUTPUT)
    build_workbook().save(out_path)
    print(f"Report saved: {out_path}")
